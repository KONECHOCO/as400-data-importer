from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, APIRouter, HTTPException, Depends, BackgroundTasks, UploadFile, File, Form, WebSocket, WebSocketDisconnect, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from starlette.staticfiles import StaticFiles
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta, timedelta
from passlib.context import CryptContext
import jwt
import secrets
from bson import ObjectId
import paypalrestsdk
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
import asyncio
import json

# ── LOGGING ──────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).parent

# ── MONGODB ──────────────────────────────────────────────────
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client_db = AsyncIOMotorClient(mongo_url)
db = client_db[os.environ.get('DB_NAME', 'as400importer')]

# ── JWT ──────────────────────────────────────────────────────
JWT_SECRET = os.environ.get("JWT_SECRET", secrets.token_hex(32))
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# ── ENCRYPTION KEY INITIALIZATION (AES-256) ──────────────────
from cryptography.fernet import Fernet
ENCRYPTION_KEY = os.environ.get("ENCRYPTION_KEY", "")
if not ENCRYPTION_KEY:
    env_path = ROOT_DIR / ".env"
    if env_path.exists():
        with open(env_path, "r") as f:
            for line in f:
                if line.startswith("ENCRYPTION_KEY="):
                    ENCRYPTION_KEY = line.split("=", 1)[1].strip()
                    break
    if not ENCRYPTION_KEY:
        ENCRYPTION_KEY = Fernet.generate_key().decode()
        try:
            mode = "a" if env_path.exists() else "w"
            with open(env_path, mode) as f:
                f.write(f"\nENCRYPTION_KEY={ENCRYPTION_KEY}\n")
            logger.info("Generata nuova ENCRYPTION_KEY e salvata in .env")
        except Exception as e:
            logger.warning(f"Impossibile salvare ENCRYPTION_KEY in .env: {e}")

def encrypt_password(password: str) -> str:
    if not password:
        return ""
    f = Fernet(ENCRYPTION_KEY.encode())
    return f"enc:{f.encrypt(password.encode()).decode()}"

def decrypt_password(encrypted_str: str) -> str:
    if not encrypted_str:
        return ""
    if encrypted_str.startswith("enc:"):
        try:
            f = Fernet(ENCRYPTION_KEY.encode())
            return f.decrypt(encrypted_str[4:].encode()).decode()
        except Exception as e:
            logger.error(f"Errore decifratura Fernet: {e}")
            raise Exception("Impossibile decifrare la password con la chiave attuale")
    import base64
    try:
        return base64.b64decode(encrypted_str.encode()).decode()
    except Exception as e:
        logger.error(f"Errore decodifica Base64 fallback: {e}")
        return encrypted_str

async def get_decrypted_password_and_migrate(conn_doc: dict) -> str:
    pwd_enc = conn_doc.get("password_enc", "")
    if not pwd_enc:
        return ""
    if pwd_enc.startswith("enc:") or pwd_enc.startswith("gAAAA"):
        token = pwd_enc[4:] if pwd_enc.startswith("enc:") else pwd_enc
        try:
            f = Fernet(ENCRYPTION_KEY.encode())
            return f.decrypt(token.encode()).decode()
        except Exception as e:
            logger.error(f"Errore decifratura Fernet per connessione {conn_doc.get('_id')}: {e}")
            raise HTTPException(status_code=500, detail="Chiave di cifratura non valida o password corrotta.")
    else:
        import base64
        try:
            decrypted = base64.b64decode(pwd_enc.encode()).decode()
            new_pwd_enc = encrypt_password(decrypted)
            await db.connections.update_one(
                {"_id": conn_doc["_id"]},
                {"$set": {"password_enc": new_pwd_enc}}
            )
            logger.info(f"Connessione {conn_doc.get('_id')} migrata con successo a cifratura AES-256")
            return decrypted
        except Exception as e:
            logger.error(f"Errore migrazione/decodifica Base64 per connessione {conn_doc.get('_id')}: {e}")
            return pwd_enc


def is_select_query(sql: str) -> bool:
    import re
    no_comments = re.sub(r'--.*$', '', sql, flags=re.MULTILINE)
    no_comments = re.sub(r'/\*.*?\*/', '', no_comments, flags=re.DOTALL)
    cleaned = no_comments.strip().upper()
    return cleaned.startswith("SELECT") or cleaned.startswith("WITH")

async def log_audit(user: dict, action: str, connection_id: str, details: str, rows_count: int):
    try:
        conn = None
        if connection_id:
            try:
                conn = await db.connections.find_one({"_id": ObjectId(connection_id)})
            except:
                pass
        
        audit_entry = {
            "user_id": str(user.get("_id")),
            "user_email": user.get("email"),
            "user_name": user.get("name"),
            "company": user.get("company", ""),
            "action": action,
            "connection_name": conn.get("name") if conn else "Sconosciuta",
            "connection_host": conn.get("host") if conn else "N/A",
            "details": details,
            "rows_count": rows_count,
            "created_at": datetime.now(timezone.utc)
        }
        await db.audit_logs.insert_one(audit_entry)
        logger.info(f"Audit log registrato: {action} da {user.get('email')}")
    except Exception as e:
        logger.error(f"Errore scrittura audit log: {e}")

# ── EMAIL (Sendgrid) ─────────────────────────────────────────
SENDGRID_API_KEY = os.environ.get("SENDGRID_API_KEY", "")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "AS400 Importer <noreply@ikonetsolutions.com>")

# ── PAYPAL ────────────────────────────────────────────────────
PAYPAL_CLIENT_ID  = os.environ.get("PAYPAL_CLIENT_ID", "sb")
PAYPAL_SECRET     = os.environ.get("PAYPAL_SECRET", "sb")
PAYPAL_MODE       = os.environ.get("PAYPAL_MODE", "sandbox")
FRONTEND_URL      = os.environ.get("FRONTEND_URL", "https://as400.ikonetsolutions.com")

paypalrestsdk.configure({
    "mode": PAYPAL_MODE,
    "client_id": PAYPAL_CLIENT_ID,
    "client_secret": PAYPAL_SECRET
})

# ── PIANI ABBONAMENTO ─────────────────────────────────────────
PLANS = {
    "starter": {
        "id": "starter",
        "name": "Starter",
        "price": 29.00,
        "currency": "EUR",
        "interval": "month",
        "description": "Per piccole aziende",
        "features": [
            "1 connessione AS/400",
            "Import fino a 50.000 righe/mese",
            "Export CSV ed Excel",
            "1 utente",
            "Support email"
        ],
        "limits": {
            "connections": 1,
            "users": 1,
            "rows_per_month": 50000,
            "export_formats": ["csv", "xlsx"],
            "saved_queries": 10,
            "history_days": 30
        }
    },
    "business": {
        "id": "business",
        "name": "Business",
        "price": 79.00,
        "currency": "EUR",
        "interval": "month",
        "description": "Per aziende in crescita",
        "popular": True,
        "features": [
            "3 connessioni AS/400",
            "Import illimitato",
            "Export CSV, Excel, XML, JSON",
            "5 utenti",
            "Saved queries illimitate",
            "Storico 90 giorni",
            "Support prioritario"
        ],
        "limits": {
            "connections": 3,
            "users": 5,
            "rows_per_month": -1,
            "export_formats": ["csv", "xlsx", "xml", "json"],
            "saved_queries": -1,
            "history_days": 90
        }
    },
    "enterprise": {
        "id": "enterprise",
        "name": "Enterprise",
        "price": 199.00,
        "currency": "EUR",
        "interval": "month",
        "description": "Per grandi organizzazioni",
        "features": [
            "Connessioni illimitate",
            "Utenti illimitati",
            "Tutte le funzionalità",
            "Storico illimitato",
            "API REST dedicata",
            "SLA garantito",
            "Support dedicato"
        ],
        "limits": {
            "connections": -1,
            "users": -1,
            "rows_per_month": -1,
            "export_formats": ["csv", "xlsx", "xml", "json"],
            "saved_queries": -1,
            "history_days": -1
        }
    }
}

# ── APP ───────────────────────────────────────────────────────
app = FastAPI(title="AS400 Data Importer API", version="2.0.0")
api_router = APIRouter(prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

security = HTTPBearer()

# ── AGENT WEBSOCKET CANAL ────────────────────────────────────
active_agents: Dict[str, WebSocket] = {}
pending_requests: Dict[str, asyncio.Event] = {}
request_responses: Dict[str, Any] = {}

async def get_user_license_key(user: dict) -> Optional[str]:
    lic = await db.licenses.find_one({"user_id": str(user.get("id")), "revoked": {"$ne": True}})
    if not lic:
        lic = await db.licenses.find_one({"user_id": str(user.get("_id")), "revoked": {"$ne": True}})
    return lic.get("license_key") if lic else None

async def send_agent_command(license_key: str, action: str, data: dict, timeout: float = 30.0) -> dict:
    if license_key not in active_agents:
        raise HTTPException(status_code=503, detail="L'agente locale per questa licenza non è attualmente connesso.")
    ws = active_agents[license_key]
    req_id = str(uuid.uuid4())
    event = asyncio.Event()
    pending_requests[req_id] = event
    payload = {"action": action, "request_id": req_id, "data": data}
    try:
        await ws.send_text(json.dumps(payload))
        await asyncio.wait_for(event.wait(), timeout=timeout)
        response = request_responses.pop(req_id, {})
        return response
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="Timeout: l'agente locale non ha risposto in tempo.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore di comunicazione con l'agente: {e}")
    finally:
        pending_requests.pop(req_id, None)
        request_responses.pop(req_id, None)

@app.websocket("/api/ws/agent")
async def ws_agent(websocket: WebSocket, license_key: str):
    await websocket.accept()
    lic = await db.licenses.find_one({"license_key": license_key})
    if not lic or lic.get("revoked") or (lic.get("expires_at") and lic.get("expires_at").replace(tzinfo=timezone.utc) < datetime.now(timezone.utc)):
        await websocket.close(code=4003, reason="Licenza non valida o scaduta")
        return
    if license_key in active_agents:
        try:
            await active_agents[license_key].close()
        except:
            pass
    active_agents[license_key] = websocket
    logger.info(f"Agente connesso con licenza: {license_key}")
    try:
        while True:
            msg_str = await websocket.receive_text()
            try:
                msg = json.loads(msg_str)
                req_id = msg.get("request_id")
                if req_id in pending_requests:
                    request_responses[req_id] = msg
                    pending_requests[req_id].set()
            except Exception as e:
                logger.error(f"Errore parsing risposta agente: {e}")
    except WebSocketDisconnect:
        logger.info(f"Agente disconnesso con licenza: {license_key}")
    finally:
        if active_agents.get(license_key) == websocket:
            del active_agents[license_key]

# ══════════════════════════════════════════════════════════════
#  MODELS
# ══════════════════════════════════════════════════════════════

class UserRegister(BaseModel):
    email: EmailStr
    password: str
    name: str
    company: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    company: str
    role: str
    plan: str
    plan_status: str
    plan_expires: Optional[datetime]
    trial_ends: Optional[datetime]
    is_active: bool
    created_at: datetime

class AS400Connection(BaseModel):
    name: str
    host: str
    user: str
    password: str
    port: int = 446
    library: str = "*LIBL"
    description: Optional[str] = None
    is_agent_mediated: Optional[bool] = False

class AS400ConnectionUpdate(BaseModel):
    name: str
    host: str
    user: str
    password: Optional[str] = None
    port: int = 446
    library: str = "*LIBL"
    description: Optional[str] = None
    is_agent_mediated: Optional[bool] = False

class AS400ConnectionResponse(BaseModel):
    id: str
    name: str
    host: str
    user: str
    port: int
    library: str
    description: Optional[str]
    created_at: datetime
    last_used: Optional[datetime]
    is_agent_mediated: bool = False

class ImportRequest(BaseModel):
    connection_id: str
    library: str
    table_name: str
    mode: str  # create, insert, update
    field_config: Optional[List[Dict]] = None

class QueryRequest(BaseModel):
    connection_id: str
    sql: str
    limit: Optional[int] = 1000

class SavedQuery(BaseModel):
    name: str
    sql: str
    connection_id: Optional[str] = None
    tags: Optional[str] = ""

class SubscriptionCreate(BaseModel):
    plan_id: str

class PayPalCaptureRequest(BaseModel):
    payment_id: str
    payer_id: str
    plan_id: str

class ExportColumnConfig(BaseModel):
    original: str
    inc: bool
    outputHeader: str
    textFormat: str

class ExportPrepareRequest(BaseModel):
    connection_id: str
    sql: str
    format: str
    columns_config: List[ExportColumnConfig]
    send_email: Optional[bool] = False

class AdminUserResponse(BaseModel):
    id: str
    mongo_id: str
    email: str
    name: str
    company: str
    role: str
    plan: str
    plan_status: str
    plan_expires: Optional[datetime] = None
    trial_ends: Optional[datetime] = None
    is_active: bool
    is_admin: bool
    created_at: datetime

class AdminUserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    company: str
    role: Optional[str] = "user"
    is_admin: Optional[bool] = False
    plan: Optional[str] = "starter"
    plan_status: Optional[str] = "trial"
    trial_ends: Optional[datetime] = None
    plan_expires: Optional[datetime] = None
    is_active: Optional[bool] = True

class AdminUserUpdate(BaseModel):
    email: Optional[EmailStr] = None
    password: Optional[str] = None
    name: Optional[str] = None
    company: Optional[str] = None
    role: Optional[str] = None
    is_admin: Optional[bool] = None
    plan: Optional[str] = None
    plan_status: Optional[str] = None
    trial_ends: Optional[datetime] = None
    plan_expires: Optional[datetime] = None
    is_active: Optional[bool] = None

class AdminLicenseUpdate(BaseModel):
    user_id: Optional[str] = None
    plan: Optional[str] = None
    expires_at: Optional[datetime] = None
    max_connections: Optional[int] = None
    notes: Optional[str] = None
    features: Optional[List[str]] = None
    revoked: Optional[bool] = None

# ══════════════════════════════════════════════════════════════
#  AUTH HELPERS
# ══════════════════════════════════════════════════════════════

import bcrypt

def hash_password(password: str) -> str:
    pwd_bytes = password.encode('utf-8')[:72]
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(pwd_bytes, salt).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    if not password or not hashed:
        return False
    # 1. Direct bcrypt check
    try:
        if hashed.startswith("$2b$") or hashed.startswith("$2a$") or hashed.startswith("$2y$"):
            return bcrypt.checkpw(password.encode('utf-8')[:72], hashed.encode('utf-8'))
    except Exception as e:
        logger.error(f"Errore bcrypt check: {e}")
    # 2. Linux crypt check for legacy $5$ hashes
    try:
        import crypt
        if crypt.crypt(password, hashed) == hashed:
            return True
    except Exception:
        pass
    # 3. Fallback passlib
    try:
        from passlib.context import CryptContext
        pwd_ctx = CryptContext(schemes=["bcrypt", "sha256_crypt"], deprecated="auto")
        if pwd_ctx.verify(password, hashed):
            return True
    except Exception:
        pass
    return False

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Utente non trovato")
        if not user.get("is_active", True):
            raise HTTPException(status_code=401, detail="Account disabilitato")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token scaduto")
    except Exception:
        raise HTTPException(status_code=401, detail="Token non valido")

async def require_active_plan(user = Depends(get_current_user)):
    """Verifica che l'utente abbia un piano attivo o trial"""
    plan_status = user.get("plan_status", "trial")
    if plan_status == "expired":
        raise HTTPException(status_code=403, detail="Abbonamento scaduto. Rinnova il piano per continuare.")
    if plan_status == "cancelled":
        raise HTTPException(status_code=403, detail="Abbonamento cancellato.")
    return user

def check_plan_limit(user: dict, limit_key: str, current_count: int = 0) -> bool:
    """Controlla se l'utente ha raggiunto il limite del piano"""
    plan_id = user.get("plan", "starter")
    plan = PLANS.get(plan_id, PLANS["starter"])
    limit = plan["limits"].get(limit_key, 0)
    if limit == -1:  # illimitato
        return True
    return current_count < limit

# ══════════════════════════════════════════════════════════════
#  AUTH ENDPOINTS
# ══════════════════════════════════════════════════════════════

@api_router.post("/auth/register")
async def register(data: UserRegister, background_tasks: BackgroundTasks):
    # Verifica email già usata
    existing = await db.users.find_one({"email": data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email già registrata")

    # Crea utente con trial 14 giorni
    user_id = str(uuid.uuid4())
    trial_ends = datetime.now(timezone.utc) + timedelta(days=14)
    user = {
        "_id": ObjectId(),
        "id": user_id,
        "email": data.email.lower(),
        "password": hash_password(data.password),
        "name": data.name,
        "company": data.company,
        "role": "admin",
        "plan": "starter",
        "plan_status": "trial",  # trial, active, expired, cancelled
        "plan_expires": None,
        "trial_ends": trial_ends,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    await db.users.insert_one(user)

    # Genera licenza di prova (trial) per 14 giorni
    license_key = generate_license_key(user_id)
    license_doc = {
        "license_key": license_key,
        "user_id": user_id,
        "plan": "starter",
        "expires_at": trial_ends,
        "max_connections": 1,
        "notes": "Trial license generated automatically on registration",
        "revoked": False,
        "hardware_id": None,
        "hostname": None,
        "suspicious_attempts": 0,
        "features": ["import", "export", "query"],
        "created_at": datetime.now(timezone.utc)
    }
    await db.licenses.insert_one(license_doc)

    # Email di benvenuto con licenza e download link
    background_tasks.add_task(send_welcome_email, data.email, data.name, trial_ends, license_key)

    token = create_token(str(user["_id"]), user["email"], user["role"])
    return {
        "token": token,
        "user": {
            "id": str(user["_id"]),
            "email": user["email"],
            "name": user["name"],
            "company": user["company"],
            "role": user["role"],
            "plan": user["plan"],
            "plan_status": user["plan_status"],
            "trial_ends": user["trial_ends"].isoformat()
        },
        "message": f"Benvenuto! Hai 14 giorni di prova gratuita."
    }

@api_router.post("/auth/login")
async def login(data: UserLogin):
    user = await db.users.find_one({"email": data.email.lower()})
    if not user or not verify_password(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Email o password non corretti")
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="Account disabilitato")

    # Controlla se trial è scaduto
    trial_ends = user.get("trial_ends")
    if user.get("plan_status") == "trial" and trial_ends:
        if datetime.now(timezone.utc) > trial_ends.replace(tzinfo=timezone.utc):
            await db.users.update_one(
                {"_id": user["_id"]},
                {"$set": {"plan_status": "expired"}}
            )
            user["plan_status"] = "expired"

    token = create_token(str(user["_id"]), user["email"], user["role"])
    return {
        "token": token,
        "user": {
            "id": str(user["_id"]),
            "email": user["email"],
            "name": user["name"],
            "company": user.get("company", ""),
            "role": user["role"],
            "plan": user.get("plan", "starter"),
            "plan_status": user.get("plan_status", "trial"),
            "trial_ends": user.get("trial_ends", "").isoformat() if user.get("trial_ends") else None,
            "plan_expires": user.get("plan_expires", "").isoformat() if user.get("plan_expires") else None
        }
    }

@api_router.get("/auth/me")
async def get_me(user = Depends(get_current_user)):
    return {
        "id": str(user["_id"]),
        "email": user["email"],
        "name": user["name"],
        "company": user.get("company", ""),
        "role": user["role"],
        "plan": user.get("plan", "starter"),
        "plan_status": user.get("plan_status", "trial"),
        "trial_ends": user.get("trial_ends", "").isoformat() if user.get("trial_ends") else None,
        "plan_expires": user.get("plan_expires", "").isoformat() if user.get("plan_expires") else None,
        "is_admin": user.get("is_admin", False)
    }

@api_router.post("/auth/change-password")
async def change_password(data: dict, user = Depends(get_current_user)):
    old_pwd = data.get("old_password", "")
    new_pwd = data.get("new_password", "")
    if not verify_password(old_pwd, user["password"]):
        raise HTTPException(status_code=400, detail="Password attuale non corretta")
    await db.users.update_one(
        {"_id": user["_id"]},
        {"$set": {"password": hash_password(new_pwd)}}
    )
    return {"message": "Password aggiornata"}

# ══════════════════════════════════════════════════════════════
#  AS400 CONNECTIONS
# ══════════════════════════════════════════════════════════════

@api_router.get("/connections")
async def get_connections(user = Depends(require_active_plan)):
    conns = await db.connections.find(
        {"user_id": str(user["_id"]), "is_deleted": {"$ne": True}}
    ).to_list(100)
    return [{
        "id": str(c["_id"]),
        "name": c["name"],
        "host": c["host"],
        "user": c["user"],
        "port": c.get("port", 446),
        "library": c.get("library", "*LIBL"),
        "description": c.get("description"),
        "created_at": c["created_at"].isoformat(),
        "last_used": c.get("last_used", "").isoformat() if c.get("last_used") else None,
        "is_agent_mediated": c.get("is_agent_mediated", False)
    } for c in conns]

@api_router.post("/connections")
async def create_connection(data: AS400Connection, user = Depends(require_active_plan)):
    # Controlla limite connessioni del piano
    count = await db.connections.count_documents(
        {"user_id": str(user["_id"]), "is_deleted": {"$ne": True}}
    )
    if not check_plan_limit(user, "connections", count):
        plan = PLANS.get(user.get("plan", "starter")) or PLANS["starter"]
        max_conn = plan["limits"]["connections"]
        raise HTTPException(
            status_code=403,
            detail=f"Limite connessioni raggiunto ({max_conn}). Aggiorna il piano."
        )

    # Cripta la password con AES-256
    encrypted_pwd = encrypt_password(data.password)

    conn = {
        "user_id": str(user["_id"]),
        "name": data.name,
        "host": data.host,
        "user": data.user,
        "password_enc": encrypted_pwd,
        "port": data.port,
        "library": data.library,
        "description": data.description,
        "created_at": datetime.now(timezone.utc),
        "last_used": None,
        "is_deleted": False,
        "is_agent_mediated": getattr(data, "is_agent_mediated", False)
    }
    result = await db.connections.insert_one(conn)
    return {"id": str(result.inserted_id), "message": "Connessione creata"}

@api_router.delete("/connections/{conn_id}")
async def delete_connection(conn_id: str, user = Depends(get_current_user)):
    await db.connections.update_one(
        {"_id": ObjectId(conn_id), "user_id": str(user["_id"])},
        {"$set": {"is_deleted": True}}
    )
    return {"message": "Connessione eliminata"}

@api_router.put("/connections/{conn_id}")
async def update_connection(conn_id: str, data: AS400ConnectionUpdate, user = Depends(require_active_plan)):
    conn = await db.connections.find_one(
        {"_id": ObjectId(conn_id), "user_id": str(user["_id"]), "is_deleted": {"$ne": True}}
    )
    if not conn:
        raise HTTPException(status_code=404, detail="Connessione non trovata")
    
    update_data = {
        "name": data.name,
        "host": data.host,
        "user": data.user,
        "port": data.port,
        "library": data.library,
        "description": data.description,
        "is_agent_mediated": getattr(data, "is_agent_mediated", False)
    }
    
    if data.password and data.password.strip():
        update_data["password_enc"] = encrypt_password(data.password)
        
    await db.connections.update_one(
        {"_id": ObjectId(conn_id)},
        {"$set": update_data}
    )
    
    await log_audit(
        user,
        "edit-connection",
        conn_id,
        f"Modificata connessione '{data.name}' ({data.host})",
        0
    )
    
    return {"message": "Connessione aggiornata"}

@api_router.post("/connections/{conn_id}/test")
async def test_connection(conn_id: str, user = Depends(require_active_plan)):
    conn = await db.connections.find_one(
        {"_id": ObjectId(conn_id), "user_id": str(user["_id"])}
    )
    if not conn:
        raise HTTPException(status_code=404, detail="Connessione non trovata")

    try:
        password = await get_decrypted_password_and_migrate(conn)
        if conn.get("is_agent_mediated"):
            license_key = await get_user_license_key(user)
            if not license_key:
                raise HTTPException(status_code=400, detail="Licenza non trovata per questo utente.")
            result = await send_agent_command(license_key, "test-connection", {
                "host": conn["host"],
                "user": conn["user"],
                "password": password
            })
        else:
            result = await asyncio.get_event_loop().run_in_executor(
                None, _test_as400_connection, conn["host"], conn["user"], password
            )
        if result.get("success"):
            await db.connections.update_one(
                {"_id": conn["_id"]},
                {"$set": {"last_used": datetime.now(timezone.utc)}}
            )
        # Log audit
        await log_audit(
            user,
            "test-connection",
            conn_id,
            f"Test connessione '{conn.get('name')}' ({conn.get('host')}) - Risultato: {'Successo' if result.get('success') else 'Fallito'}",
            0
        )
        return result
    except Exception as e:
        return {"success": False, "message": str(e)}

@api_router.get("/connections/{conn_id}/schemas")
async def get_connection_schemas(conn_id: str, user = Depends(require_active_plan)):
    conn = await db.connections.find_one({"_id": ObjectId(conn_id), "user_id": str(user["_id"])})
    if not conn:
        raise HTTPException(status_code=404, detail="Connessione non trovata")
    
    password = await get_decrypted_password_and_migrate(conn)
    sql = "SELECT DISTINCT SCHEMA_NAME FROM QSYS2.SYSSCHEMAS ORDER BY SCHEMA_NAME"
    
    try:
        if conn.get("is_agent_mediated"):
            license_key = await get_user_license_key(user)
            res = await send_agent_command(license_key, "query", {
                "connection": {"host": conn["host"], "user": conn["user"], "password": password, "port": conn.get("port", 446)},
                "sql": sql, "limit": 1000
            })
            if not res.get("success"):
                sql_fb = "SELECT DISTINCT TABLE_SCHEMA AS SCHEMA_NAME FROM SYSIBM.TABLES ORDER BY TABLE_SCHEMA"
                res = await send_agent_command(license_key, "query", {
                    "connection": {"host": conn["host"], "user": conn["user"], "password": password, "port": conn.get("port", 446)},
                    "sql": sql_fb, "limit": 1000
                })
            rows = res.get("rows", [])
            schemas = [r[0] for r in rows]
        else:
            try:
                res = _sync_query(conn["host"], conn["user"], password, sql, 1000)
            except:
                sql_fb = "SELECT DISTINCT TABLE_SCHEMA AS SCHEMA_NAME FROM SYSIBM.TABLES ORDER BY TABLE_SCHEMA"
                res = _sync_query(conn["host"], conn["user"], password, sql_fb, 1000)
            schemas = [r["SCHEMA_NAME"] for r in res]
        return {"schemas": schemas}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore caricamento schemi: {e}")

@api_router.get("/connections/{conn_id}/schemas/{schema}/tables")
async def get_connection_tables(conn_id: str, schema: str, user = Depends(require_active_plan)):
    conn = await db.connections.find_one({"_id": ObjectId(conn_id), "user_id": str(user["_id"])})
    if not conn:
        raise HTTPException(status_code=404, detail="Connessione non trovata")
    
    password = await get_decrypted_password_and_migrate(conn)
    sql = f"SELECT TABLE_NAME, COALESCE(TABLE_TEXT, '') AS TABLE_TEXT FROM QSYS2.SYSTABLES WHERE TABLE_SCHEMA = '{schema.upper()}' ORDER BY TABLE_NAME"
    
    try:
        if conn.get("is_agent_mediated"):
            license_key = await get_user_license_key(user)
            res = await send_agent_command(license_key, "query", {
                "connection": {"host": conn["host"], "user": conn["user"], "password": password, "port": conn.get("port", 446)},
                "sql": sql, "limit": 1000
            })
            if not res.get("success"):
                sql_fb = f"SELECT TABLE_NAME, '' AS TABLE_TEXT FROM SYSIBM.TABLES WHERE TABLE_SCHEMA = '{schema.upper()}' ORDER BY TABLE_NAME"
                res = await send_agent_command(license_key, "query", {
                    "connection": {"host": conn["host"], "user": conn["user"], "password": password, "port": conn.get("port", 446)},
                    "sql": sql_fb, "limit": 1000
                })
            rows = res.get("rows", [])
            tables = [{"name": r[0], "description": r[1]} for r in rows]
        else:
            try:
                res = _sync_query(conn["host"], conn["user"], password, sql, 1000)
            except:
                sql_fb = f"SELECT TABLE_NAME, '' AS TABLE_TEXT FROM SYSIBM.TABLES WHERE TABLE_SCHEMA = '{schema.upper()}' ORDER BY TABLE_NAME"
                res = _sync_query(conn["host"], conn["user"], password, sql_fb, 1000)
            tables = [{"name": r["TABLE_NAME"], "description": r["TABLE_TEXT"]} for r in res]
        return {"tables": tables}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore caricamento tabelle: {e}")

@api_router.get("/connections/{conn_id}/schemas/{schema}/tables/{table}/columns")
async def get_connection_columns(conn_id: str, schema: str, table: str, user = Depends(require_active_plan)):
    conn = await db.connections.find_one({"_id": ObjectId(conn_id), "user_id": str(user["_id"])})
    if not conn:
        raise HTTPException(status_code=404, detail="Connessione non trovata")
    
    password = await get_decrypted_password_and_migrate(conn)
    sql = f"SELECT COLUMN_NAME, DATA_TYPE, LENGTH, COALESCE(COLUMN_TEXT, '') AS COLUMN_TEXT FROM QSYS2.SYSCOLUMNS WHERE TABLE_SCHEMA = '{schema.upper()}' AND TABLE_NAME = '{table.upper()}' ORDER BY ORDINAL_POSITION"
    
    try:
        if conn.get("is_agent_mediated"):
            license_key = await get_user_license_key(user)
            res = await send_agent_command(license_key, "query", {
                "connection": {"host": conn["host"], "user": conn["user"], "password": password, "port": conn.get("port", 446)},
                "sql": sql, "limit": 1000
            })
            if not res.get("success"):
                sql_fb = f"SELECT COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH AS LENGTH, '' AS COLUMN_TEXT FROM SYSIBM.COLUMNS WHERE TABLE_SCHEMA = '{schema.upper()}' AND TABLE_NAME = '{table.upper()}'"
                res = await send_agent_command(license_key, "query", {
                    "connection": {"host": conn["host"], "user": conn["user"], "password": password, "port": conn.get("port", 446)},
                    "sql": sql_fb, "limit": 1000
                })
            rows = res.get("rows", [])
            columns = [{"name": r[0], "type": r[1], "length": r[2], "description": r[3]} for r in rows]
        else:
            try:
                res = _sync_query(conn["host"], conn["user"], password, sql, 1000)
            except:
                sql_fb = f"SELECT COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH AS LENGTH, '' AS COLUMN_TEXT FROM SYSIBM.COLUMNS WHERE TABLE_SCHEMA = '{schema.upper()}' AND TABLE_NAME = '{table.upper()}'"
                res = _sync_query(conn["host"], conn["user"], password, sql_fb, 1000)
            columns = [{"name": r["COLUMN_NAME"], "type": r["DATA_TYPE"], "length": r["LENGTH"], "description": r["COLUMN_TEXT"]} for r in res]
        return {"columns": columns}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore caricamento colonne: {e}")

def _test_as400_connection(host: str, user: str, password: str) -> dict:
    """Test connessione AS/400 sincrona (eseguita in thread)"""
    try:
        import jaydebeapi
        jar_path = os.path.join(ROOT_DIR, "lib", "jt400.jar")
        url = (f"jdbc:as400://{host};"
               f"user={user};password={password};"
               f"prompt=false;secure=false")
        conn = jaydebeapi.connect(
            "com.ibm.as400.access.AS400JDBCDriver",
            url, [user, password], jar_path
        )
        conn.close()
        return {"success": True, "message": "Connessione riuscita"}
    except Exception as e:
        return {"success": False, "message": str(e)}

def smart_read_df(source, ext="csv"):
    """
    Lettore intelligente per file Excel, XML, CSV e TXT:
    - Rilevamento automatico della codifica (utf-8-sig, utf-8, latin-1)
    - Rilevamento automatico del delimitatore (|, ;, \t, ,, :)
    - Rilevamento automatico dell'intestazione (header vs no-header)
    """
    import pandas as pd
    import io
    import csv

    ext = str(ext).lower().lstrip(".")
    if ext in ["xlsx", "xls"]:
        if isinstance(source, bytes):
            return pd.read_excel(io.BytesIO(source))
        return pd.read_excel(source)
    elif ext == "xml":
        if isinstance(source, bytes):
            return pd.read_xml(io.BytesIO(source))
        return pd.read_xml(source)

    # File TXT o CSV
    if isinstance(source, bytes):
        content_bytes = source
    else:
        with open(source, "rb") as f:
            content_bytes = f.read()

    # 1. Rileva codifica
    text_sample = None
    encoding_used = "utf-8-sig"
    for enc in ["utf-8-sig", "utf-8", "latin-1"]:
        try:
            text_sample = content_bytes[:100000].decode(enc)
            encoding_used = enc
            break
        except UnicodeDecodeError:
            continue
    if text_sample is None:
        text_sample = content_bytes[:100000].decode("latin-1", errors="replace")
        encoding_used = "latin-1"

    # 2. Rileva delimitatore
    sample_lines = [line for line in text_sample.splitlines()[:30] if line.strip()]
    delims = ['|', ';', '\t', ',', ':']
    counts = {d: [line.count(d) for line in sample_lines] for d in delims}

    best_delim = ','
    best_score = 0
    for d, line_counts in counts.items():
        if not line_counts:
            continue
        avg_c = sum(line_counts) / len(line_counts)
        if avg_c > 0:
            consistent = len(set(line_counts)) == 1
            score = avg_c * (10 if consistent else 1)
            if score > best_score:
                best_score = score
                best_delim = d

    # 3. Rileva intestazione (header)
    header = 0
    try:
        sample_for_sniffer = "\n".join(sample_lines)
        if not csv.Sniffer().has_header(sample_for_sniffer):
            header = None
    except Exception:
        header = None

    df = pd.read_csv(
        io.BytesIO(content_bytes),
        sep=best_delim,
        header=header,
        encoding=encoding_used,
        dtype=str,
        on_bad_lines="skip"
    )

    if header is None:
        df.columns = [f"COL_{i+1}" for i in range(len(df.columns))]

    return df

# ══════════════════════════════════════════════════════════════
#  IMPORT
# ══════════════════════════════════════════════════════════════

@api_router.post("/import/preview")
async def import_preview(file: UploadFile = File(...), user = Depends(require_active_plan)):
    try:
        content = await file.read()
        filename = file.filename
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "csv"
        
        import pandas as pd
        
        # Leggi dataframe con smart_read_df
        df = smart_read_df(content, ext)
            
        columns = list(df.columns)
        df_preview = df.head(5).fillna("")
        preview_data = df_preview.values.tolist()
        
        # Rileva tipi suggeriti
        suggested_types = {}
        for col in columns:
            series = df[col].dropna()
            if series.empty:
                suggested_types[col] = "VARCHAR(255)"
                continue
            if pd.api.types.is_integer_dtype(series):
                suggested_types[col] = "INTEGER"
            elif pd.api.types.is_numeric_dtype(series):
                suggested_types[col] = "DECIMAL(15,4)"
            elif pd.api.types.is_datetime64_any_dtype(series):
                suggested_types[col] = "TIMESTAMP"
            else:
                max_len = int(series.astype(str).str.len().max())
                buffer_len = max(50, min(1000, (max_len // 50 + 1) * 50))
                suggested_types[col] = f"VARCHAR({buffer_len})"
                
        return {
            "columns": columns,
            "preview": preview_data,
            "suggested_types": suggested_types
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Errore lettura anteprima file: {e}")

TEMP_IMPORT_FILES = {}

@api_router.get("/agent/download-import-file/{file_token}")
async def download_import_file(file_token: str):
    if file_token not in TEMP_IMPORT_FILES:
        raise HTTPException(status_code=404, detail="File non trovato o scaduto")
    data = TEMP_IMPORT_FILES[file_token]
    from fastapi.responses import Response
    return Response(content=data["content"], media_type="application/octet-stream")

@api_router.post("/import")
async def import_data(
    file: UploadFile = File(...),
    connection_id: str = Form(...),
    library: str = Form(...),
    table_name: str = Form(...),
    mode: str = Form(...),
    field_config: Optional[str] = Form(None),
    user = Depends(require_active_plan),
    background_tasks: BackgroundTasks = BackgroundTasks()
):
    import json
    request = ImportRequest(
        connection_id=connection_id,
        library=library,
        table_name=table_name,
        mode=mode,
        field_config=json.loads(field_config) if field_config else None
    )
    # Controlla limite righe mensili
    plan_id = user.get("plan", "starter")
    plan = PLANS.get(plan_id, PLANS["starter"])
    max_rows = plan["limits"]["rows_per_month"]

    if max_rows != -1:
        # Conta righe importate questo mese
        month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0)
        rows_this_month = await db.operations.aggregate([
            {"$match": {
                "user_id": str(user["_id"]),
                "type": "import",
                "created_at": {"$gte": month_start}
            }},
            {"$group": {"_id": None, "total": {"$sum": "$rows_count"}}}
        ]).to_list(1)
        current_rows = rows_this_month[0]["total"] if rows_this_month else 0
        if current_rows >= max_rows:
            raise HTTPException(
                status_code=403,
                detail=f"Limite righe mensile raggiunto ({max_rows:,}). Aggiorna il piano."
            )

    # Leggi file
    content = await file.read()
    filename = file.filename
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "csv"

    # Log operazione
    op_id = str(uuid.uuid4())
    await db.operations.insert_one({
        "id": op_id,
        "user_id": str(user["_id"]),
        "type": "import",
        "status": "pending",
        "filename": filename,
        "connection_id": request.connection_id,
        "library": request.library,
        "table_name": request.table_name,
        "mode": request.mode,
        "rows_count": 0,
        "created_at": datetime.now(timezone.utc)
    })

    # Esegui import in background
    background_tasks.add_task(
        _run_import, op_id, content, ext, request, user, plan
    )

    return {"operation_id": op_id, "message": "Import avviato", "status": "pending"}

async def _run_import(op_id, content, ext, request, user, plan):
    """Esegue l'import in background"""
    try:
        import pandas as pd
        import io

        # Leggi dataframe
        df = smart_read_df(content, ext)

        rows = len(df)

        # Aggiorna contatore
        await db.operations.update_one(
            {"id": op_id},
            {"$set": {"rows_count": rows, "status": "running"}}
        )

        # Ottieni connessione
        conn_doc = await db.connections.find_one({"_id": ObjectId(request.connection_id)})
        if not conn_doc:
            raise Exception("Connessione non trovata")

        password = await get_decrypted_password_and_migrate(conn_doc)

        if conn_doc.get("is_agent_mediated"):
            file_token = str(uuid.uuid4())
            TEMP_IMPORT_FILES[file_token] = {
                "content": content,
                "ext": ext,
                "created_at": time.time()
            }
            # Pulisci file vecchi più di 30 minuti
            now = time.time()
            for k in list(TEMP_IMPORT_FILES.keys()):
                if now - TEMP_IMPORT_FILES[k]["created_at"] > 1800:
                    TEMP_IMPORT_FILES.pop(k, None)

            file_url = f"https://as400.ikonetsolutions.com/api/agent/download-import-file/{file_token}"

            license_key = await get_user_license_key(user)
            if not license_key:
                raise Exception("Licenza non trovata per questo utente.")
            result = await send_agent_command(license_key, "import", {
                "connection": {
                    "host": conn_doc["host"],
                    "user": conn_doc["user"],
                    "password": password,
                    "port": conn_doc.get("port", 446)
                },
                "library": request.library,
                "table_name": request.table_name,
                "mode": request.mode,
                "file_url": file_url,
                "file_ext": ext,
                "field_config": request.field_config
            }, timeout=600.0)
        else:
            # Import AS/400 in thread
            result = await asyncio.get_event_loop().run_in_executor(
                None, _sync_import,
                conn_doc["host"], conn_doc["user"], password,
                request.library, request.table_name, request.mode,
                df, request.field_config
            )

        await db.operations.update_one(
            {"id": op_id},
            {"$set": {
                "status": "completed" if result["success"] else "failed",
                "result": result,
                "error": result.get("error") if not result["success"] else None,
                "completed_at": datetime.now(timezone.utc)
            }}
        )
        # Log audit per completamento
        op_doc = await db.operations.find_one({"id": op_id})
        filename = op_doc.get("filename", "unknown")
        details = f"Import tabella {request.library}.{request.table_name} ({request.mode}) da file '{filename}' - Risultato: {'Completato' if result.get('success') else 'Fallito'}"
        await log_audit(
            user,
            "import",
            request.connection_id,
            details,
            result.get("inserted", 0) if result.get("success") else 0
        )

    except Exception as e:
        await db.operations.update_one(
            {"id": op_id},
            {"$set": {"status": "failed", "error": str(e)}}
        )
        # Log audit per errore prima del completamento
        details = f"Import tabella {request.library}.{request.table_name} ({request.mode}) fallito: {e}"
        await log_audit(
            user,
            "import-error",
            request.connection_id,
            details,
            0
        )

def _sync_import(host, user, password, library, table_name, mode, df, field_config):
    """Import AS/400 sincrono"""
    try:
        import jaydebeapi
        jar_path = os.path.join(ROOT_DIR, "lib", "jt400.jar")
        url = (f"jdbc:as400://{host};user={user};password={password};"
               f"prompt=false;secure=false")
        conn = jaydebeapi.connect(
            "com.ibm.as400.access.AS400JDBCDriver",
            url, [user, password], jar_path
        )
        cursor = conn.cursor()

        # Prepara colonne
        if field_config:
            columns = [fc["as400_name"] for fc in field_config]
            orig_cols = [fc["original"] for fc in field_config]
            data_rows = [tuple(row) for row in df[orig_cols].values]
            sql_types = {fc["as400_name"]: fc["sql_type"] for fc in field_config}
        else:
            columns = list(df.columns)
            data_rows = [tuple(row) for row in df.values]
            sql_types = {col: "VARCHAR(500)" for col in columns}

        # Crea tabella se necessario
        if mode == "create":
            cols_sql = ", ".join([f"{col} {sql_types.get(col, 'VARCHAR(500)')}" for col in columns])
            try:
                cursor.execute(f"CREATE TABLE {library}.{table_name} ({cols_sql})")
                conn.commit()
            except Exception as create_err:
                err_msg = str(create_err).lower()
                if "already exists" not in err_msg and "sql0601" not in err_msg and "duplicate" not in err_msg:
                    raise Exception(f"Errore creazione tabella {library}.{table_name}: {create_err}")

        # Insert/Update
        inserted = 0
        errors = []
        if mode in ["create", "insert"]:
            cols_str = ", ".join(columns)
            placeholders = ", ".join(["?" for _ in columns])
            sql = f"INSERT INTO {library}.{table_name} ({cols_str}) VALUES ({placeholders})"
            for i in range(0, len(data_rows), 500):
                batch = data_rows[i:i+500]
                batch_vals = []
                for row in batch:
                    vals = [None if (v is None or (isinstance(v, float) and v != v)) else v for v in row]
                    batch_vals.append(vals)
                try:
                    cursor.executemany(sql, batch_vals)
                    inserted += len(batch)
                except Exception as batch_err:
                    logger.warning(f"Batch insert fallito nel server cloud, fallback riga per riga: {batch_err}")
                    for vals in batch_vals:
                        try:
                            cursor.execute(sql, vals)
                            inserted += 1
                        except Exception as row_err:
                            errors.append(str(row_err))
                conn.commit()

        cursor.close()
        conn.close()
        if errors:
            return {"success": False, "error": f"Errore import: {errors[0]}", "inserted": inserted, "total": len(data_rows), "errors": errors[:5]}
        return {"success": True, "inserted": inserted, "total": len(data_rows)}

    except Exception as e:
        import traceback
        return {"success": False, "error": str(e), "detail": traceback.format_exc()}
@api_router.get("/import/{op_id}/status")
async def get_import_status(op_id: str, user = Depends(get_current_user)):
    op = await db.operations.find_one({"id": op_id, "user_id": str(user["_id"])})
    if not op:
        raise HTTPException(status_code=404, detail="Operazione non trovata")
    return {
        "id": op["id"],
        "status": op["status"],
        "rows_count": op.get("rows_count", 0),
        "result": op.get("result"),
        "error": op.get("error"),
        "created_at": op["created_at"].isoformat()
    }

# ══════════════════════════════════════════════════════════════
#  EXPORT / QUERY
# ══════════════════════════════════════════════════════════════

@api_router.post("/query")
async def execute_query(data: QueryRequest, user = Depends(require_active_plan)):
    # Valida query in sola lettura
    if not is_select_query(data.sql):
        raise HTTPException(
            status_code=400,
            detail="Operazione non consentita. Sono permesse solo query di lettura (SELECT o WITH)."
        )

    conn_doc = await db.connections.find_one(
        {"_id": ObjectId(data.connection_id), "user_id": str(user["_id"])}
    )
    if not conn_doc:
        raise HTTPException(status_code=404, detail="Connessione non trovata")

    # Controlla formato export nel piano
    password = await get_decrypted_password_and_migrate(conn_doc)

    try:
        if conn_doc.get("is_agent_mediated"):
            license_key = await get_user_license_key(user)
            if not license_key:
                raise HTTPException(status_code=400, detail="Licenza non trovata per questo utente.")
            agent_res = await send_agent_command(license_key, "query", {
                "connection": {
                    "host": conn_doc["host"],
                    "user": conn_doc["user"],
                    "password": password,
                    "port": conn_doc.get("port", 446)
                },
                "sql": data.sql,
                "limit": data.limit
            }, timeout=60.0)
            if not agent_res.get("success"):
                raise Exception(agent_res.get("error", "Errore sconosciuto dell'agente"))
            cols = agent_res.get("columns", [])
            rows_data = agent_res.get("rows", [])
            result = [dict(zip(cols, [str(v) if v is not None else "" for v in row])) for row in rows_data]
        else:
            result = await asyncio.get_event_loop().run_in_executor(
                None, _sync_query, conn_doc["host"], conn_doc["user"], password, data.sql, data.limit
            )

        # Log operazione export
        await db.operations.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": str(user["_id"]),
            "type": "export",
            "status": "completed",
            "sql": data.sql,
            "rows_count": len(result),
            "created_at": datetime.now(timezone.utc)
        })

        # Log audit
        await log_audit(
            user,
            "query",
            data.connection_id,
            f"Query SQL eseguita: {data.sql}",
            len(result)
        )

        return {"data": result, "rows": len(result)}
    except Exception as e:
        await log_audit(
            user,
            "query-error",
            data.connection_id,
            f"Errore query SQL ({data.sql}): {e}",
            0
        )
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/export/prepare")
async def export_prepare(data: ExportPrepareRequest, user = Depends(require_active_plan)):
    token = str(uuid.uuid4())
    expiry = datetime.now(timezone.utc) + timedelta(minutes=10)
    
    await db.export_prepares.insert_one({
        "token": token,
        "user_id": str(user["_id"]),
        "connection_id": data.connection_id,
        "sql": data.sql,
        "format": data.format,
        "columns_config": [c.model_dump() for c in data.columns_config],
        "send_email": data.send_email,
        "expires_at": expiry
    })
    return {"token": token}

@api_router.get("/export/download/{token}")
async def export_download(token: str, background_tasks: BackgroundTasks):
    prep = await db.export_prepares.find_one({"token": token})
    if not prep:
        raise HTTPException(status_code=404, detail="Token di download non valido o scaduto.")
        
    expires_at = prep.get("expires_at")
    if expires_at and expires_at.replace(tzinfo=timezone.utc) < datetime.now(timezone.utc):
        raise HTTPException(status_code=410, detail="Token scaduto. Rigenera il file.")
        
    user_id = prep.get("user_id")
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="Utente non trovato.")
        
    connection_id = prep.get("connection_id")
    sql = prep.get("sql")
    fmt = prep.get("format", "xlsx")
    columns_config = prep.get("columns_config", [])
    send_email = prep.get("send_email", False)
    
    conn_doc = await db.connections.find_one({
        "_id": ObjectId(connection_id), "user_id": user_id
    })
    if not conn_doc:
        raise HTTPException(status_code=404, detail="Connessione non trovata")
        
    password = await get_decrypted_password_and_migrate(conn_doc)
    
    try:
        if conn_doc.get("is_agent_mediated"):
            license_key = await get_user_license_key(user)
            if not license_key:
                raise Exception("Licenza non trovata per questo utente.")
            agent_res = await send_agent_command(license_key, "query", {
                "connection": {
                    "host": conn_doc["host"],
                    "user": conn_doc["user"],
                    "password": password,
                    "port": conn_doc.get("port", 446)
                },
                "sql": sql,
                "limit": None
            }, timeout=300.0)
            if not agent_res.get("success"):
                raise Exception(agent_res.get("error", "Errore sconosciuto dell'agente"))
            cols = agent_res.get("columns", [])
            rows_data = agent_res.get("rows", [])
            raw_data = [dict(zip(cols, [v if v is not None else "" for v in row])) for row in rows_data]
        else:
            raw_data = await asyncio.get_event_loop().run_in_executor(
                None, _sync_query, conn_doc["host"], conn_doc["user"], password, sql, None
            )
    except Exception as e:
        logger.error(f"Errore esecuzione query per export: {e}")
        raise HTTPException(status_code=500, detail=f"Errore nel recupero dei dati da AS/400: {e}")
        
    if not raw_data:
        raise HTTPException(status_code=400, detail="La query non ha restituito alcun dato.")
        
    import pandas as pd
    df = pd.DataFrame(raw_data)
    
    if columns_config:
        inc_configs = [c for c in columns_config if c.get("inc", True)]
        columns_to_keep = [c.get("original") for c in inc_configs if c.get("original") in df.columns]
        
        if columns_to_keep:
            df = df[columns_to_keep]
            rename_dict = {}
            for c in inc_configs:
                orig = c.get("original")
                if orig not in df.columns:
                    continue
                out_header = c.get("outputHeader", orig)
                txt_format = c.get("textFormat", "nessuno")
                
                if txt_format == "maiuscolo":
                    df[orig] = df[orig].astype(str).str.upper()
                elif txt_format == "minuscolo":
                    df[orig] = df[orig].astype(str).str.lower()
                    
                rename_dict[orig] = out_header
                
            df = df.rename(columns=rename_dict)
            
    try:
        content_type = "application/octet-stream"
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{fmt}"
        
        if fmt == "xlsx":
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            file_bytes = _generate_xlsx_bytes(df)
        elif fmt == "csv":
            content_type = "text/csv"
            file_bytes = df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
        elif fmt == "tsv":
            content_type = "text/tab-separated-values"
            file_bytes = df.to_csv(index=False, sep="\t", encoding="utf-8-sig").encode("utf-8-sig")
        elif fmt == "json":
            content_type = "application/json"
            file_bytes = df.to_json(orient="records", indent=2).encode("utf-8")
        elif fmt == "xml":
            content_type = "application/xml"
            file_bytes = _generate_xml_bytes(df)
        elif fmt == "pdf":
            content_type = "application/pdf"
            file_bytes = _generate_pdf_bytes(df, filename)
        else:
            raise Exception(f"Formato non supportato: {fmt}")
            
    except Exception as e:
        logger.error(f"Errore generazione file export: {e}")
        raise HTTPException(status_code=500, detail=f"Errore nella generazione del file di export: {e}")
        
    if send_email:
        background_tasks.add_task(
            _send_export_email_task, user.get("email"), user.get("name"), filename, file_bytes, content_type
        )
        
    await log_audit(
        user,
        "export-download",
        connection_id,
        f"Export file '{filename}' ({fmt}) scaricato con successo. Righe: {len(df)}",
        len(df)
    )
    
    headers = {
        "Content-Disposition": f"attachment; filename=\"{filename}\"",
        "Access-Control-Expose-Headers": "Content-Disposition"
    }
    return Response(content=file_bytes, media_type=content_type, headers=headers)

def _generate_xlsx_bytes(df) -> bytes:
    import io
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Convert numeric columns where possible
    for col in df.columns:
        try:
            df[col] = pd.to_numeric(df[col])
        except:
            pass

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Export")
        
        # Get workbook and sheet
        workbook = writer.book
        worksheet = writer.sheets["Export"]
        
        # Enable grid lines visible
        worksheet.views.sheetView[0].showGridLines = True
        
        # Define styles
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_alignment = Alignment(horizontal='left', vertical='center')
        
        cell_font = Font(name='Segoe UI', size=10)
        
        thin_side = Side(style='thin', color='E2E8F0')
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Style the headers
        worksheet.row_dimensions[1].height = 26
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border
            
        # Style the data rows and auto-fit column widths
        col_widths = {}
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = len(str(col_name))
            col_widths[col_idx] = max_len
            
        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 20
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.alignment = Alignment(
                    horizontal='right' if cell.data_type in ['n', 'f'] else 'left', 
                    vertical='center'
                )
                cell.border = cell_border
                
                # Check value length for auto-width
                val_str = str(cell.value) if cell.value is not None else ""
                col_widths[col_idx] = max(col_widths[col_idx], len(val_str))
                
        # Set column widths
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = max(width + 4, 12)

    buffer.seek(0)
    return buffer.getvalue()

def _generate_xml_bytes(df) -> bytes:
    import io
    import xml.etree.ElementTree as ET
    root = ET.Element("export")
    for idx, row in df.iterrows():
        row_el = ET.SubElement(root, "row")
        for col in df.columns:
            clean_col = "".join([c if c.isalnum() else "_" for c in str(col)])
            if not clean_col or not clean_col[0].isalpha():
                clean_col = "col_" + clean_col
            col_el = ET.SubElement(row_el, clean_col)
            col_el.text = str(row[col]) if row[col] is not None else ""
    buffer = io.BytesIO()
    tree = ET.ElementTree(root)
    tree.write(buffer, encoding="utf-8", xml_declaration=True)
    return buffer.getvalue()

def _generate_pdf_bytes(df, filename: str) -> bytes:
    import io
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    buffer = io.BytesIO()
    pagesize = landscape(letter) if len(df.columns) > 5 else letter
    doc = SimpleDocTemplate(buffer, pagesize=pagesize, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=12,
        leading=14,
        textColor=colors.HexColor("#1e293b"),
        spaceAfter=10
    )
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=7,
        leading=9
    )
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor("#ffffff")
    )
    
    story.append(Paragraph(f"Export Dati - {filename}", title_style))
    story.append(Spacer(1, 10))
    
    pdf_df = df.head(500)
    data = []
    data.append([Paragraph(str(col), header_style) for col in pdf_df.columns])
    
    for idx, row in pdf_df.iterrows():
        data.append([Paragraph(str(val) if val is not None else "", cell_style) for val in row])
        
    col_width = (doc.width) / len(pdf_df.columns)
    table = Table(data, colWidths=[col_width] * len(pdf_df.columns))
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1e293b")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    
    story.append(table)
    
    if len(df) > 500:
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"* Nota: Mostrate le prime 500 righe di {len(df)} totali nel PDF.", cell_style))
        
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

async def _send_export_email_task(email: str, name: str, filename: str, file_bytes: bytes, content_type: str):
    if not SENDGRID_API_KEY:
        logger.warning("Nessuna chiave API email configurata (Sendgrid). Invio email saltato.")
        return
        
    import base64
    file_b64 = base64.b64encode(file_bytes).decode()
    
    subject = f"Il tuo file di esportazione è pronto: {filename}"
    html = f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:40px;background:#0f1117;border-radius:12px;color:#ffffff;">
<h1 style="color:#4f9cf9;text-align:center;">Esportazione Completata! 📊</h1>
<p style="color:#9ca3af;">Ciao {name},</p>
<p style="color:#9ca3af;">Il file richiesto <strong>{filename}</strong> è pronto ed è allegato a questa email.</p>
<p style="color:#4b5563;font-size:12px;text-align:center;margin-top:40px;">© 2026 Ikonet Solutions</p>
</div>"""
            
    if SENDGRID_API_KEY:
        try:
            logger.info("Tentativo invio email con Sendgrid...")
            from sendgrid import SendGridAPIClient
            from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
            sg = SendGridAPIClient(SENDGRID_API_KEY)
            
            message = Mail(
                from_email=SENDER_EMAIL,
                to_emails=email,
                subject=subject,
                html_content=html
            )
            
            attached_file = Attachment(
                FileContent(file_b64),
                FileName(filename),
                FileType(content_type),
                Disposition('attachment')
            )
            message.attachment = attached_file
            
            sg.send(message)
            logger.info(f"Email di export inviata via SendGrid a {email}")
        except Exception as e:
            logger.error(f"Errore invio email con Sendgrid: {e}")

def _sync_query(host, user, password, sql, limit):
    """Esegue query AS/400 sincrona"""
    import jaydebeapi
    jar_path = os.path.join(ROOT_DIR, "lib", "jt400.jar")
    url = (f"jdbc:as400://{host};user={user};password={password};"
           f"prompt=false;secure=false")
    conn = jaydebeapi.connect(
        "com.ibm.as400.access.AS400JDBCDriver",
        url, [user, password], jar_path
    )
    cursor = conn.cursor()

    # Aggiungi FETCH FIRST se non presente
    if limit and "FETCH FIRST" not in sql.upper():
        sql = f"{sql} FETCH FIRST {limit} ROWS ONLY"

    cursor.execute(sql)
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    return [dict(zip(columns, [str(v) if v is not None else "" for v in row])) for row in rows]

@api_router.get("/audit-logs")
async def get_audit_logs(user = Depends(get_current_user)):
    # Solo amministratori
    if user.get("role") != "admin" and not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Accesso negato. Solo gli amministratori possono accedere al registro di audit.")
    
    query = {}
    if not user.get("is_admin"):
        query["company"] = user.get("company", "")
        
    logs = await db.audit_logs.find(query).sort("created_at", -1).limit(300).to_list(300)
    return [{
        "id": str(l["_id"]),
        "user_email": l["user_email"],
        "user_name": l["user_name"],
        "company": l.get("company", ""),
        "action": l["action"],
        "connection_name": l.get("connection_name", "N/A"),
        "connection_host": l.get("connection_host", ""),
        "details": l["details"],
        "rows_count": l.get("rows_count", 0),
        "created_at": l["created_at"].isoformat()
    } for l in logs]

# ══════════════════════════════════════════════════════════════
#  SAVED QUERIES
# ══════════════════════════════════════════════════════════════

@api_router.get("/saved-queries")
async def get_saved_queries(user = Depends(get_current_user)):
    queries = await db.saved_queries.find(
        {"user_id": str(user["_id"])}
    ).sort("created_at", -1).to_list(200)
    return [{
        "id": str(q["_id"]),
        "name": q["name"],
        "sql": q["sql"],
        "tags": q.get("tags", ""),
        "uses": q.get("uses", 0),
        "created_at": q["created_at"].isoformat()
    } for q in queries]

@api_router.post("/saved-queries")
async def save_query(data: SavedQuery, user = Depends(require_active_plan)):
    # Controlla limite saved queries
    count = await db.saved_queries.count_documents({"user_id": str(user["_id"])})
    if not check_plan_limit(user, "saved_queries", count):
        raise HTTPException(status_code=403, detail="Limite saved queries raggiunto. Aggiorna il piano.")

    result = await db.saved_queries.insert_one({
        "user_id": str(user["_id"]),
        "name": data.name,
        "sql": data.sql,
        "tags": data.tags,
        "uses": 0,
        "created_at": datetime.now(timezone.utc)
    })
    return {"id": str(result.inserted_id), "message": "Query salvata"}

@api_router.delete("/saved-queries/{query_id}")
async def delete_saved_query(query_id: str, user = Depends(get_current_user)):
    await db.saved_queries.delete_one({"_id": ObjectId(query_id), "user_id": str(user["_id"])})
    return {"message": "Query eliminata"}

# ══════════════════════════════════════════════════════════════
#  STORICO OPERAZIONI
# ══════════════════════════════════════════════════════════════

@api_router.get("/operations")
async def get_operations(user = Depends(get_current_user)):
    plan_id = user.get("plan", "starter")
    plan = PLANS.get(plan_id, PLANS["starter"])
    history_days = plan["limits"]["history_days"]

    query = {"user_id": str(user["_id"])}
    if history_days != -1:
        cutoff = datetime.now(timezone.utc) - timedelta(days=history_days)
        query["created_at"] = {"$gte": cutoff}

    ops = await db.operations.find(query).sort("created_at", -1).limit(100).to_list(100)
    return [{
        "id": op.get("id"),
        "type": op["type"],
        "status": op["status"],
        "filename": op.get("filename"),
        "sql": op.get("sql"),
        "rows_count": op.get("rows_count", 0),
        "library": op.get("library"),
        "table_name": op.get("table_name"),
        "created_at": op["created_at"].isoformat()
    } for op in ops]

# ══════════════════════════════════════════════════════════════
#  PIANI E ABBONAMENTI
# ══════════════════════════════════════════════════════════════

@api_router.get("/plans")
async def get_plans():
    return {"plans": list(PLANS.values()), "paypal_client_id": PAYPAL_CLIENT_ID}

@api_router.get("/subscription")
async def get_subscription(user = Depends(get_current_user)):
    plan_id = user.get("plan", "starter")
    plan = PLANS.get(plan_id, PLANS["starter"])
    return {
        "plan": plan,
        "plan_status": user.get("plan_status", "trial"),
        "plan_expires": user.get("plan_expires", "").isoformat() if user.get("plan_expires") else None,
        "trial_ends": user.get("trial_ends", "").isoformat() if user.get("trial_ends") else None,
    }

@api_router.post("/subscription/create-payment")
async def create_subscription_payment(data: SubscriptionCreate, user = Depends(get_current_user)):
    plan = PLANS.get(data.plan_id)
    if not plan:
        raise HTTPException(status_code=400, detail="Piano non valido")

    payment = paypalrestsdk.Payment({
        "intent": "sale",
        "payer": {"payment_method": "paypal"},
        "redirect_urls": {
            "return_url": f"{FRONTEND_URL}/payment/success?plan={data.plan_id}",
            "cancel_url": f"{FRONTEND_URL}/payment/cancel"
        },
        "transactions": [{
            "amount": {
                "total": f"{plan['price']:.2f}",
                "currency": plan["currency"]
            },
            "description": f"AS400 Importer — Piano {plan['name']} ({plan['interval']})"
        }]
    })

    if payment.create():
        approval_url = next(l["href"] for l in payment.links if l["rel"] == "approval_url")
        return {"payment_id": payment.id, "approval_url": approval_url}
    else:
        raise HTTPException(status_code=500, detail=f"Errore PayPal: {payment.error}")

@api_router.post("/subscription/capture-payment")
async def capture_subscription_payment(
    data: PayPalCaptureRequest,
    user = Depends(get_current_user),
    background_tasks: BackgroundTasks = BackgroundTasks()
):
    plan = PLANS.get(data.plan_id)
    if not plan:
        raise HTTPException(status_code=400, detail="Piano non valido")

    payment = paypalrestsdk.Payment.find(data.payment_id)
    if payment.execute({"payer_id": data.payer_id}):
        # Aggiorna abbonamento utente (30 giorni)
        expires = datetime.now(timezone.utc) + timedelta(days=30)
        await db.users.update_one(
            {"_id": user["_id"]},
            {"$set": {
                "plan": data.plan_id,
                "plan_status": "active",
                "plan_expires": expires,
                "trial_ends": None
            }}
        )

        # Salva transazione
        await db.transactions.insert_one({
            "user_id": str(user["_id"]),
            "payment_id": data.payment_id,
            "plan_id": data.plan_id,
            "amount": plan["price"],
            "currency": plan["currency"],
            "status": "completed",
            "created_at": datetime.now(timezone.utc)
        })

        # Email conferma
        background_tasks.add_task(
            send_subscription_email, user["email"], user["name"], plan, expires
        )

        return {"success": True, "message": f"Piano {plan['name']} attivato!", "expires": expires.isoformat()}
    else:
        raise HTTPException(status_code=500, detail=f"Errore cattura PayPal: {payment.error}")

# ══════════════════════════════════════════════════════════════
#  EMAIL HELPERS
# ══════════════════════════════════════════════════════════════

async def send_welcome_email(email: str, name: str, trial_ends: datetime, license_key: str):
    if not SENDGRID_API_KEY:
        logger.warning("Nessuna chiave API email configurata (Sendgrid). Welcome email non inviata.")
        return

    subject = "Benvenuto su AS400 Data Importer — 14 giorni gratis!"
    html_content = f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:40px;background:#0d0f1a;border-radius:16px;color:#fff;border:1px solid #1e293b;">
  <!-- Header -->
  <div style="margin-bottom:30px;border-bottom:1px solid #1e293b;padding-bottom:20px;">
    <div style="width:36px;height:36px;background:linear-gradient(to bottom right, #3b82f6, #10b981);border-radius:8px;display:inline-block;text-align:center;line-height:36px;font-weight:900;color:white;font-size:16px;margin-right:12px;vertical-align:middle;">IK</div>
    <div style="display:inline-block;vertical-align:middle;">
      <div style="color:white;font-weight:700;font-size:16px;line-height:1.2;margin:0;">ikonet</div>
      <div style="color:#60a5fa;font-size:10px;letter-spacing:2px;font-weight:700;margin:0;">AS400 DATA IMPORTER</div>
    </div>
  </div>

  <!-- Body -->
  <h1 style="font-size:24px;font-weight:700;color:#fff;margin-top:0;margin-bottom:15px;">Benvenuto, {name}! 👋</h1>
  <p style="color:#94a3b8;font-size:15px;line-height:1.6;margin-bottom:25px;">Il tuo account è attivo su <strong>AS400 Data Importer</strong>. Segui i 2 passi qui sotto per iniziare subito.</p>

  <!-- Step 1 -->
  <div style="margin-bottom:30px;background:#151824;padding:20px;border-radius:12px;border:1px solid #1e293b;">
    <h2 style="font-size:16px;font-weight:700;color:#3b82f6;margin-top:0;margin-bottom:10px;">
      Scarica e installa l'applicazione
    </h2>
    <p style="color:#94a3b8;font-size:14px;line-height:1.5;margin-bottom:20px;">Clicca il pulsante per scaricare l'installer dell'Agente per Windows (circa 100 MB). Durante l'installazione, inserisci la chiave di licenza riportata sotto quando richiesto.</p>
    <div style="text-align:center;margin-bottom:15px;">
      <a href="{FRONTEND_URL}/AS400AgentSetup.exe" style="background:#2563eb;color:white;padding:12px 24px;border-radius:8px;text-decoration:none;font-weight:700;font-size:14px;display:inline-block;border:1px solid #3b82f6;box-shadow:0 4px 6px -1px rgba(0,0,0,0.1);">⬇ Scarica Installer Agente (.exe)</a>
    </div>
    <div style="color:#64748b;font-size:11px;text-align:center;">Windows 10/11 (64-bit) — Tutto incluso (Java JRE 17 integrato, non richiede installazioni esterne)</div>
  </div>

  <!-- Step 2 -->
  <div style="margin-bottom:30px;background:#151824;padding:20px;border-radius:12px;border:1px solid #1e293b;">
    <h2 style="font-size:16px;font-weight:700;color:#10b981;margin-top:0;margin-bottom:10px;">
      La tua licenza di prova
    </h2>
    <p style="color:#94a3b8;font-size:14px;line-height:1.5;margin-bottom:15px;">Copia questa chiave di licenza di prova gratuita (valida per 14 giorni) e incollala nella procedura d'installazione guidata:</p>
    <div style="background:#0d0f1a;border:1px dashed #10b981;color:#34d399;padding:14px;border-radius:8px;font-family:monospace;font-size:18px;text-align:center;letter-spacing:1px;font-weight:700;margin:15px 0;">
      {license_key}
    </div>
    <div style="color:#64748b;font-size:11px;text-align:center;">Questo codice di prova scade il {trial_ends.strftime('%d/%m/%Y')}.</div>
  </div>

  <!-- Footer -->
  <div style="border-top:1px solid #1e293b;padding-top:20px;text-align:center;">
    <p style="color:#94a3b8;font-size:13px;line-height:1.5;">Per qualsiasi dubbio o supporto, rispondi a questa email.</p>
    <p style="color:#475569;font-size:11px;margin-top:20px;">© 2026 Ikonet Solutions</p>
  </div>
</div>"""

    if SENDGRID_API_KEY:
        try:
            logger.info("Tentativo invio welcome email con SendGrid...")
            from sendgrid import SendGridAPIClient
            from sendgrid.helpers.mail import Mail
            sg = SendGridAPIClient(SENDGRID_API_KEY)
            msg = Mail(
                from_email=SENDER_EMAIL,
                to_emails=email,
                subject=subject,
                html_content=html_content
            )
            sg.send(msg)
            logger.info(f"Welcome email inviata via SendGrid a {email}")
        except Exception as e:
            logger.error(f"Errore welcome email SendGrid: {e}")


async def send_subscription_email(email: str, name: str, plan: dict, expires: datetime):
    if not SENDGRID_API_KEY:
        return
    try:
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import Mail
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        msg = Mail(
            from_email="supporto@ikonetsolutions.com",
            to_emails=email,
            subject=f"Piano {plan['name']} attivato — AS400 Data Importer",
            html_content=f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:40px;background:#0f1117;border-radius:12px;">
<h1 style="color:#22c55e;text-align:center;">🎉 Abbonamento Attivato!</h1>
<p style="color:#9ca3af;">Grazie <strong style="color:#fff;">{name}</strong>! Piano <strong style="color:#22c55e;">{plan['name']}</strong> attivato.</p>
<p style="color:#9ca3af;">💰 {plan['price']:.2f} {plan['currency']}/mese — Scade il {expires.strftime('%d/%m/%Y')}</p>
<div style="text-align:center;margin:30px 0;">
<a href="{FRONTEND_URL}/dashboard" style="background:#22c55e;color:white;padding:14px 32px;border-radius:10px;text-decoration:none;font-weight:700;">Vai alla Dashboard →</a>
</div>
<p style="color:#4b5563;font-size:12px;text-align:center;">© 2026 Ikonet Solutions</p>
</div>"""
        )
        sg.send(msg)
        logger.info(f"Subscription email inviata a {email}")
    except Exception as e:
        logger.error(f"Errore subscription email: {e}")


async def health():
    return {"status": "ok", "service": "AS400 Data Importer API", "version": "2.0.0"}


@api_router.post("/auth/forgot-password")
async def forgot_password(data: dict, background_tasks: BackgroundTasks):
    email = data.get("email", "").lower()
    user = await db.users.find_one({"email": email})
    if user and SENDGRID_API_KEY:
        token = secrets.token_urlsafe(32)
        await db.users.update_one(
            {"_id": user["_id"]},
            {"$set": {"reset_token": token, "reset_expires": datetime.now(timezone.utc) + timedelta(hours=2)}}
        )
        reset_url = f"{FRONTEND_URL}/reset-password?token={token}"
        background_tasks.add_task(send_reset_email, email, user["name"], reset_url)
    # Rispondi sempre OK per sicurezza
    return {"message": "Se l'email esiste, riceverai le istruzioni"}

async def send_reset_email(email: str, name: str, reset_url: str):
    if not SENDGRID_API_KEY:
        logger.warning("SENDGRID_API_KEY non configurata")
        return
    try:
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import Mail
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        message = Mail(
            from_email="supporto@ikonetsolutions.com",
            to_emails=email,
            subject="Reset password — AS400 Data Importer",
            html_content=f"""<h2>Ciao {name}!</h2>
            <p>Hai richiesto il reset della password per AS400 Data Importer.</p>
            <br>
            <a href="{reset_url}" style="background:#4f9cf9;color:white;padding:12px 24px;border-radius:8px;text-decoration:none;">Reset Password</a>
            <br><br>
            <p style="color:#666;font-size:12px;">Il link scade tra 2 ore.</p>
            <p>2026 Ikonet Solutions</p>"""
        )
        sg.send(message)
        logger.info(f"Reset email inviata a {email}")
    except Exception as e:
        logger.error(f"Errore reset email SendGrid: {e}")



@api_router.post("/auth/reset-password")
async def reset_password(data: dict):
    token = data.get("token", "")
    password = data.get("password", "")
    logger.info(f"Reset attempt with token: {token[:20]}...")
    if not token or not password:
        raise HTTPException(status_code=400, detail="Token e password richiesti")
    user = await db.users.find_one({"reset_token": token})
    logger.info(f"User found: {user is not None}")
    if not user:
        raise HTTPException(status_code=400, detail="Token non valido o scaduto")
    expires = user.get("reset_expires")
    if expires and datetime.now(timezone.utc) > expires.replace(tzinfo=timezone.utc):
        raise HTTPException(status_code=400, detail="Token scaduto")
    await db.users.update_one(
        {"_id": user["_id"]},
        {"$set": {"password": hash_password(password)},
         "$unset": {"reset_token": "", "reset_expires": ""}}
    )
    return {"message": "Password aggiornata con successo"}


@api_router.get("/admin/send-trial-reminders")
async def send_trial_reminders(user = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    reminder_date = now + timedelta(days=7)
    users = await db.users.find({
        "plan_status": "trial",
        "trial_ends": {"$gte": now, "$lte": reminder_date},
        "trial_reminder_sent": {"$ne": True}
    }).to_list(100)
    sent = 0
    for u in users:
        days_left = max(1, (u["trial_ends"].replace(tzinfo=timezone.utc) - now).days)
        if SENDGRID_API_KEY:
            try:
                from sendgrid import SendGridAPIClient
                from sendgrid.helpers.mail import Mail
                sg = SendGridAPIClient(SENDGRID_API_KEY)
                msg = Mail(
                    from_email="supporto@ikonetsolutions.com",
                    to_emails=u["email"],
                    subject=f"Il tuo trial scade tra {days_left} giorni — AS400 Data Importer",
                    html_content=f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:40px;background:#0f1117;border-radius:12px;">
<h1 style="color:#f59e0b;text-align:center;">⏰ Trial in scadenza!</h1>
<p style="color:#9ca3af;">Ciao <strong style="color:#fff;">{u['name']}</strong>, il tuo trial scade tra <strong style="color:#f59e0b;">{days_left} giorni</strong>.</p>
<div style="text-align:center;margin:30px 0;">
<a href="{FRONTEND_URL}/plans" style="background:#4f9cf9;color:white;padding:14px 32px;border-radius:10px;text-decoration:none;font-weight:700;">Scegli il piano →</a>
</div>
<p style="color:#4b5563;font-size:12px;text-align:center;">© 2026 Ikonet Solutions</p>
</div>"""
                )
                sg.send(msg)
                sent += 1
            except Exception as e:
                logger.error(f"Errore trial reminder: {e}")
        await db.users.update_one({"_id": u["_id"]}, {"$set": {"trial_reminder_sent": True}})
    return {"sent": sent}


# Forgot password endpoint (aggiunto)


# ══════════════════════════════════════════════════════════════
#  LICENSE MANAGEMENT
# ══════════════════════════════════════════════════════════════

def generate_license_key(customer_id: str) -> str:
    import hashlib, hmac as _hmac
    LICENSE_SECRET = os.environ.get("LICENSE_SECRET", "ikonet-license-secret-2024")
    raw = f"{customer_id}-{secrets.token_hex(8)}"
    h = _hmac.new(LICENSE_SECRET.encode(), raw.encode(), hashlib.sha256).hexdigest()[:12].upper()
    return f"IKONET-{h[0:4]}-{h[4:8]}-{h[8:12]}"

class LicenseActivateRequest(BaseModel):
    license_key: str
    hardware_id: str
    hostname: str
    ip: Optional[str] = None

class LicenseIssueRequest(BaseModel):
    user_id: str
    plan: str
    expires_days: int = 365
    max_connections: int = 1
    notes: Optional[str] = ""

@api_router.post("/license/verify")
async def verify_license(req: LicenseActivateRequest):
    lic = await db.licenses.find_one({"license_key": req.license_key})
    if not lic:
        raise HTTPException(status_code=404, detail="Licenza non trovata")
    now = datetime.now(timezone.utc)
    expires_at = lic.get("expires_at")
    if expires_at and expires_at.replace(tzinfo=timezone.utc) < now:
        raise HTTPException(status_code=403, detail="Licenza scaduta")
    if lic.get("revoked"):
        raise HTTPException(status_code=403, detail="Licenza revocata")
    registered_hw = lic.get("hardware_id")
    if not registered_hw:
        await db.licenses.update_one(
            {"license_key": req.license_key},
            {"$set": {"hardware_id": req.hardware_id, "hostname": req.hostname,
                      "first_activated": now, "last_seen": now, "ip": req.ip}}
        )
    elif registered_hw != req.hardware_id:
        await db.licenses.update_one(
            {"license_key": req.license_key},
            {"$inc": {"suspicious_attempts": 1}, "$set": {"last_seen": now}}
        )
        raise HTTPException(status_code=403, detail="Licenza gia attivata su un altro dispositivo")
    else:
        await db.licenses.update_one(
            {"license_key": req.license_key},
            {"$set": {"last_seen": now, "hostname": req.hostname}}
        )
    return {
        "valid": True,
        "plan": lic.get("plan", "starter"),
        "expires_at": lic.get("expires_at").isoformat() if lic.get("expires_at") else None,
        "max_connections": lic.get("max_connections", 1),
        "customer": lic.get("customer_name", ""),
        "features": lic.get("features", ["import", "export"])
    }

@api_router.get("/admin/licenses")
async def list_licenses(user=Depends(get_current_user)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Accesso negato")
    licenses = await db.licenses.find().sort("created_at", -1).to_list(200)
    for l in licenses:
        l["_id"] = str(l["_id"])
    return licenses

@api_router.post("/admin/licenses")
async def create_license(req: LicenseIssueRequest, user=Depends(get_current_user)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Accesso negato")
    key = generate_license_key(req.user_id)
    expires_at = datetime.now(timezone.utc) + timedelta(days=req.expires_days)
    doc = {
        "license_key": key,
        "user_id": req.user_id,
        "plan": req.plan,
        "expires_at": expires_at,
        "max_connections": req.max_connections,
        "notes": req.notes,
        "revoked": False,
        "hardware_id": None,
        "hostname": None,
        "suspicious_attempts": 0,
        "features": ["import", "export", "query"],
        "created_at": datetime.now(timezone.utc)
    }
    await db.licenses.insert_one(doc)
    doc["_id"] = str(doc["_id"])
    return {"license_key": key, "expires_at": expires_at.isoformat(), **doc}

@api_router.post("/admin/licenses/{license_key}/revoke")
async def revoke_license(license_key: str, user=Depends(get_current_user)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Accesso negato")
    result = await db.licenses.update_one(
        {"license_key": license_key},
        {"$set": {"revoked": True, "revoked_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Licenza non trovata")
    return {"message": "Licenza revocata"}

@api_router.post("/admin/licenses/{license_key}/reset-hardware")
async def reset_hardware(license_key: str, user=Depends(get_current_user)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Accesso negato")
    await db.licenses.update_one(
        {"license_key": license_key},
        {"$set": {"hardware_id": None, "hostname": None, "suspicious_attempts": 0}}
    )
    return {"message": "Hardware reset eseguito"}

# ── ADMIN USER CRUD ──────────────────────────────────────────

@api_router.get("/admin/users", response_model=List[AdminUserResponse])
async def admin_list_users(user=Depends(get_current_user)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Accesso negato")
    users = await db.users.find().sort("created_at", -1).to_list(1000)
    res = []
    for u in users:
        res.append(AdminUserResponse(
            id=u.get("id") or str(u["_id"]),
            mongo_id=str(u["_id"]),
            email=u["email"],
            name=u.get("name", ""),
            company=u.get("company", ""),
            role=u.get("role", "user"),
            plan=u.get("plan", "starter"),
            plan_status=u.get("plan_status", "trial"),
            plan_expires=u.get("plan_expires"),
            trial_ends=u.get("trial_ends"),
            is_active=u.get("is_active", True),
            is_admin=u.get("is_admin", False),
            created_at=u.get("created_at") or datetime.now(timezone.utc)
        ))
    return res

@api_router.post("/admin/users")
async def admin_create_user(req: AdminUserCreate, user=Depends(get_current_user)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Accesso negato")
    
    existing = await db.users.find_one({"email": req.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email già registrata")
        
    user_id = str(uuid.uuid4())
    trial_ends = req.trial_ends or (datetime.now(timezone.utc) + timedelta(days=14))
    
    new_user = {
        "_id": ObjectId(),
        "id": user_id,
        "email": req.email.lower(),
        "password": hash_password(req.password),
        "name": req.name,
        "company": req.company,
        "role": req.role,
        "plan": req.plan,
        "plan_status": req.plan_status,
        "plan_expires": req.plan_expires,
        "trial_ends": trial_ends,
        "is_active": req.is_active,
        "is_admin": req.is_admin,
        "created_at": datetime.now(timezone.utc)
    }
    await db.users.insert_one(new_user)
    
    if req.plan_status == "trial":
        key = generate_license_key(user_id)
        license_doc = {
            "license_key": key,
            "user_id": user_id,
            "plan": req.plan,
            "expires_at": trial_ends,
            "max_connections": 1,
            "notes": "Licenza di prova creata automaticamente dall'admin",
            "revoked": False,
            "hardware_id": None,
            "hostname": None,
            "suspicious_attempts": 0,
            "features": ["import", "export", "query"],
            "created_at": datetime.now(timezone.utc)
        }
        await db.licenses.insert_one(license_doc)
        
    return {"message": "Utente creato con successo", "user_id": user_id}

@api_router.put("/admin/users/{mongo_id}")
async def admin_update_user(mongo_id: str, req: AdminUserUpdate, user=Depends(get_current_user)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Accesso negato")
    
    target_user = await db.users.find_one({"_id": ObjectId(mongo_id)})
    if not target_user:
        raise HTTPException(status_code=404, detail="Utente non trovato")
        
    update_data = {}
    if req.name is not None:
        update_data["name"] = req.name
    if req.email is not None:
        existing = await db.users.find_one({"email": req.email.lower(), "_id": {"$ne": ObjectId(mongo_id)}})
        if existing:
            raise HTTPException(status_code=400, detail="Email già utilizzata da un altro utente")
        update_data["email"] = req.email.lower()
    if req.password is not None and req.password.strip() != "":
        update_data["password"] = hash_password(req.password)
    if req.company is not None:
        update_data["company"] = req.company
    if req.role is not None:
        update_data["role"] = req.role
    if req.is_admin is not None:
        update_data["is_admin"] = req.is_admin
    if req.plan is not None:
        update_data["plan"] = req.plan
    if req.plan_status is not None:
        update_data["plan_status"] = req.plan_status
    if req.trial_ends is not None:
        update_data["trial_ends"] = req.trial_ends
    if req.plan_expires is not None:
        update_data["plan_expires"] = req.plan_expires
    if req.is_active is not None:
        update_data["is_active"] = req.is_active
        
    if update_data:
        await db.users.update_one({"_id": ObjectId(mongo_id)}, {"$set": update_data})
        
    return {"message": "Utente aggiornato con successo"}

@api_router.delete("/admin/users/{mongo_id}")
async def admin_delete_user(mongo_id: str, user=Depends(get_current_user)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Accesso negato")
    
    target_user = await db.users.find_one({"_id": ObjectId(mongo_id)})
    if not target_user:
        raise HTTPException(status_code=404, detail="Utente non trovato")
        
    await db.users.delete_one({"_id": ObjectId(mongo_id)})
    
    user_uuid = target_user.get("id")
    if user_uuid:
        await db.licenses.delete_many({"user_id": user_uuid})
        await db.connections.delete_many({"user_id": user_uuid})
    
    await db.licenses.delete_many({"user_id": mongo_id})
    await db.connections.delete_many({"user_id": mongo_id})
    
    return {"message": "Utente e relative risorse eliminati con successo"}

# ── ADMIN LICENSE EXTRA CRUD ─────────────────────────────────

@api_router.put("/admin/licenses/{license_key}")
async def admin_update_license(license_key: str, req: AdminLicenseUpdate, user=Depends(get_current_user)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Accesso negato")
        
    lic = await db.licenses.find_one({"license_key": license_key})
    if not lic:
        raise HTTPException(status_code=404, detail="Licenza non trovata")
        
    update_data = {}
    if req.user_id is not None:
        update_data["user_id"] = req.user_id
    if req.plan is not None:
        update_data["plan"] = req.plan
    if req.expires_at is not None:
        update_data["expires_at"] = req.expires_at
    if req.max_connections is not None:
        update_data["max_connections"] = req.max_connections
    if req.notes is not None:
        update_data["notes"] = req.notes
    if req.features is not None:
        update_data["features"] = req.features
    if req.revoked is not None:
        update_data["revoked"] = req.revoked
        if req.revoked:
            update_data["revoked_at"] = datetime.now(timezone.utc)
            
    if update_data:
        await db.licenses.update_one({"license_key": license_key}, {"$set": update_data})
        
    return {"message": "Licenza aggiornata con successo"}

@api_router.delete("/admin/licenses/{license_key}")
async def admin_delete_license(license_key: str, user=Depends(get_current_user)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Accesso negato")
        
    result = await db.licenses.delete_one({"license_key": license_key})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Licenza non trovata")
        
    return {"message": "Licenza eliminata con successo"}

app.include_router(api_router)

# Serve frontend statico
if os.path.exists("/app/frontend/build"):
    app.mount("/", StaticFiles(directory="/app/frontend/build", html=True), name="frontend")

