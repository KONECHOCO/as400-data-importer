"""
Lettura e scrittura di tutti i formati file supportati.
Usato da /api/import/preview, /api/import e /api/export/download.
"""
import io, json, re, csv as _csv
from typing import Any

import chardet
import pandas as pd
import numpy as np

# ── Encoding ──────────────────────────────────────────────────────────────────
def detect_encoding(content: bytes) -> str:
    result = chardet.detect(content[:20_000])
    enc    = result.get("encoding") or "utf-8"
    norm   = {"ISO-8859-1": "latin-1", "windows-1252": "cp1252",
               "ascii": "utf-8", "UTF-8-SIG": "utf-8-sig"}
    return norm.get(enc, enc)

# ── CSV separator ─────────────────────────────────────────────────────────────
def detect_separator(text: str) -> str:
    sample = "\n".join(text.split("\n")[:30])
    counts = {",": 0, ";": 0, "\t": 0, "|": 0}
    for sep in counts:
        counts[sep] = sample.count(sep)
    return max(counts, key=counts.get)

# ── AS400 column name ─────────────────────────────────────────────────────────
def as400_col(name: str) -> str:
    s = re.sub(r"[^A-Z0-9]", "_", str(name).upper())
    s = re.sub(r"_+", "_", s).strip("_")
    return (s[:10] or "COL")

# ── Type detection ────────────────────────────────────────────────────────────
def detect_type(series: pd.Series) -> str:
    non_null = series.dropna().astype(str).str.strip()
    non_null = non_null[non_null != ""]
    if len(non_null) == 0:
        return "VARCHAR(256)"
    # Integer
    try:
        vals = non_null.astype(int)
        mx   = vals.abs().max()
        return "SMALLINT" if mx < 32768 else ("INTEGER" if mx < 2_147_483_648 else "BIGINT")
    except Exception:
        pass
    # Decimal
    try:
        cleaned = non_null.str.replace(",", ".", regex=False)
        cleaned.astype(float)
        if non_null.str.contains(r"[.,]\d", regex=True).any():
            return "DECIMAL(15,2)"
    except Exception:
        pass
    # Date
    try:
        pd.to_datetime(non_null, infer_datetime_format=True)
        return "DATE"
    except Exception:
        pass
    # String length
    mx = non_null.str.len().max()
    if mx <= 10:  return "VARCHAR(10)"
    if mx <= 50:  return "VARCHAR(50)"
    if mx <= 100: return "VARCHAR(100)"
    if mx <= 256: return "VARCHAR(256)"
    return "CLOB(1M)"

# ── DataFrame cleaner ─────────────────────────────────────────────────────────
def _clean(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    return df.where(pd.notna(df), None)

def _to_rows(df: pd.DataFrame) -> list[dict]:
    rows = []
    for _, row in df.iterrows():
        rows.append({c: (None if (v is None or (isinstance(v, float) and np.isnan(v))) else str(v))
                     for c, v in row.items()})
    return rows

# ── Read ──────────────────────────────────────────────────────────────────────
def read_file(content: bytes, filename: str, options: dict = None) -> tuple[pd.DataFrame, dict]:
    """
    Ritorna (DataFrame, info).
    info = {format, encoding, separator, sheets}
    """
    opt  = options or {}
    name = filename.lower()
    info = {"format": "unknown", "encoding": "utf-8", "separator": None, "sheets": []}

    hdr = opt.get("header_row", 0)
    hdr = None if hdr == -1 else hdr   # -1 → nessun header

    # Excel moderno
    if name.endswith((".xlsx", ".xlsm")):
        info["format"] = "excel"
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            info["sheets"] = wb.sheetnames
            wb.close()
        except Exception:
            pass
        sheet = opt.get("sheet", 0)
        df = pd.read_excel(io.BytesIO(content), engine="openpyxl",
                           sheet_name=sheet, header=hdr, dtype=str, keep_default_na=False)

    # Excel legacy
    elif name.endswith(".xls"):
        info["format"] = "excel_xls"
        df = pd.read_excel(io.BytesIO(content), engine="xlrd",
                           header=hdr, dtype=str, keep_default_na=False)

    # Excel binary
    elif name.endswith(".xlsb"):
        info["format"] = "excel_xlsb"
        df = pd.read_excel(io.BytesIO(content), engine="pyxlsb",
                           header=hdr, dtype=str, keep_default_na=False)

    # CSV / TXT / TSV
    elif name.endswith((".csv", ".txt", ".tsv")):
        enc  = opt.get("encoding") or detect_encoding(content)
        text = content.decode(enc, errors="replace")
        sep  = opt.get("separator") or detect_separator(text)
        info.update({"format": "csv", "encoding": enc, "separator": sep})
        df = pd.read_csv(io.StringIO(text), sep=sep, header=hdr,
                         dtype=str, keep_default_na=False, on_bad_lines="skip")

    # XML
    elif name.endswith(".xml"):
        info["format"] = "xml"
        try:
            df = pd.read_xml(io.BytesIO(content), dtype=str)
        except Exception:
            import xml.etree.ElementTree as ET
            root    = ET.parse(io.BytesIO(content)).getroot()
            records = []
            for child in root:
                rec = {}
                for elem in child:
                    rec[elem.tag] = elem.text or ""
                    for sub in elem:
                        rec[f"{elem.tag}_{sub.tag}"] = sub.text or ""
                if rec:
                    records.append(rec)
            df = pd.DataFrame(records, dtype=str)

    # JSON
    elif name.endswith(".json"):
        enc  = opt.get("encoding") or detect_encoding(content)
        text = content.decode(enc, errors="replace")
        info.update({"format": "json", "encoding": enc})
        data = json.loads(text)
        if isinstance(data, list):
            df = pd.DataFrame(data, dtype=str)
        elif isinstance(data, dict):
            for v in data.values():
                if isinstance(v, list):
                    df = pd.DataFrame(v, dtype=str); break
            else:
                df = pd.DataFrame([data], dtype=str)
        else:
            df = pd.DataFrame()

    # JSON Lines
    elif name.endswith(".jsonl"):
        enc = opt.get("encoding") or detect_encoding(content)
        info.update({"format": "jsonl", "encoding": enc})
        df = pd.read_json(io.BytesIO(content), lines=True, dtype=str)

    # Fixed Width
    elif opt.get("fixed_width") and opt.get("col_specs"):
        enc   = opt.get("encoding") or detect_encoding(content)
        text  = content.decode(enc, errors="replace")
        specs = [(s["start"], s["start"] + s["width"]) for s in opt["col_specs"]]
        names = [s["name"] for s in opt["col_specs"]]
        info.update({"format": "fixed_width", "encoding": enc})
        df = pd.read_fwf(io.StringIO(text), colspecs=specs, names=names, dtype=str)

    # Fallback CSV
    else:
        enc  = detect_encoding(content)
        text = content.decode(enc, errors="replace")
        sep  = detect_separator(text)
        info.update({"format": "csv", "encoding": enc, "separator": sep})
        df = pd.read_csv(io.StringIO(text), sep=sep, dtype=str,
                         keep_default_na=False, on_bad_lines="skip")

    return _clean(df), info


# ── Preview ───────────────────────────────────────────────────────────────────
def get_preview(content: bytes, filename: str, options: dict = None) -> dict:
    df, info = read_file(content, filename, options)
    headers  = list(df.columns)
    mapping  = [
        {"file_col": h, "as400_col": as400_col(h),
         "include": True, "type": detect_type(df[h]), "null_empty": True}
        for h in headers
    ]
    return {
        "filename":   filename,
        "format":     info["format"],
        "encoding":   info.get("encoding", "utf-8"),
        "separator":  info.get("separator"),
        "sheets":     info.get("sheets", []),
        "total_rows": len(df),
        "headers":    headers,
        "preview":    _to_rows(df.head(10)),
        "mapping":    mapping,
    }


# ── Import to AS400 ───────────────────────────────────────────────────────────
def import_df(df: pd.DataFrame, mapping: list[dict],
              open_conn, library: str, table: str, mode: str,
              op_id: str, pending: dict) -> tuple[int, list]:
    """
    open_conn: callable → JDBC connection
    Returns (inserted_rows, errors)
    """
    col_map    = {m["file_col"]: m for m in mapping if m.get("include", True)}
    file_cols  = [c for c in df.columns if c in col_map]
    if not file_cols:
        return 0, [{"row": 0, "error": "Nessuna colonna selezionata"}]

    as400_cols = [col_map[c]["as400_col"] for c in file_cols]
    types      = {col_map[c]["as400_col"]: col_map[c].get("type", "VARCHAR(256)") for c in file_cols}
    null_empty = {c: col_map[c].get("null_empty", True) for c in file_cols}

    total    = len(df)
    conn     = open_conn()
    cur      = conn.cursor()
    errors   = []
    inserted = 0

    if op_id in pending:
        pending[op_id]["total_rows"] = total

    try:
        full = f"{library.upper()}.{table.upper()}"

        if mode == "create":
            col_defs = ", ".join(f"{col} {types[col]}" for col in as400_cols)
            try:
                cur.execute(f"DROP TABLE {full}"); conn.commit()
            except Exception:
                pass
            cur.execute(f"CREATE TABLE {full} ({col_defs})"); conn.commit()

        ph  = ", ".join(["?" for _ in as400_cols])
        ins = f"INSERT INTO {full} ({', '.join(as400_cols)}) VALUES ({ph})"

        for i, (_, row) in enumerate(df.iterrows()):
            try:
                vals = []
                for fc in file_cols:
                    v = row.get(fc)
                    if v is None or (null_empty[fc] and str(v).strip() == ""):
                        vals.append(None)
                    else:
                        vals.append(str(v))
                cur.execute(ins, vals)
                inserted += 1
            except Exception as e:
                errors.append({"row": i + 2, "error": str(e)})

            if op_id in pending and i % 50 == 0:
                pending[op_id]["progress"]   = int((i + 1) / total * 100)
                pending[op_id]["rows_count"] = inserted
                pending[op_id]["rows_error"] = len(errors)

            if (i + 1) % 500 == 0:
                conn.commit()

        conn.commit()
    finally:
        cur.close(); conn.close()

    return inserted, errors


# ── Export to file ────────────────────────────────────────────────────────────
_MIME = {
    "xlsx":  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "xls":   "application/vnd.ms-excel",
    "csv":   "text/csv",
    "tsv":   "text/tab-separated-values",
    "txt":   "text/plain",
    "json":  "application/json",
    "jsonl": "application/x-ndjson",
    "xml":   "application/xml",
    "pdf":   "application/pdf",
}

def export_file(df: pd.DataFrame, fmt: str, options: dict = None) -> tuple[bytes, str]:
    """Returns (bytes, mime_type)."""
    opt = options or {}
    buf = io.BytesIO()

    if fmt == "xlsx":
        with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="Export")
            wb  = w.book
            ws  = w.sheets["Export"]
            hfmt = wb.add_format({"bold": True, "bg_color": "#1e3a5f",
                                  "font_color": "white", "border": 1})
            for ci, col in enumerate(df.columns):
                ws.write(0, ci, col, hfmt)
                width = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) else 0)
                ws.set_column(ci, ci, min(width + 2, 40))
        return buf.getvalue(), _MIME["xlsx"]

    if fmt == "xls":
        import xlwt
        wb = xlwt.Workbook()
        ws = wb.add_sheet("Export")
        hstyle = xlwt.easyxf("font: bold true; pattern: pattern solid, fore_colour dark_blue; font: colour white")
        for ci, col in enumerate(df.columns):
            ws.write(0, ci, col, hstyle)
        for ri, (_, row) in enumerate(df.iterrows(), 1):
            for ci, col in enumerate(df.columns):
                ws.write(ri, ci, row[col])
        wb.save(buf)
        return buf.getvalue(), _MIME["xls"]

    if fmt == "csv":
        sep = opt.get("separator", ",")
        enc = opt.get("encoding", "utf-8-sig")
        df.to_csv(buf, index=False, sep=sep, encoding=enc)
        return buf.getvalue(), _MIME["csv"]

    if fmt == "tsv":
        df.to_csv(buf, index=False, sep="\t", encoding="utf-8-sig")
        return buf.getvalue(), _MIME["tsv"]

    if fmt == "txt":
        df.to_csv(buf, index=False, sep=opt.get("separator", ";"), encoding="utf-8-sig")
        return buf.getvalue(), _MIME["txt"]

    if fmt == "json":
        out = df.to_json(orient="records", force_ascii=False, indent=2, date_format="iso")
        return out.encode("utf-8"), _MIME["json"]

    if fmt == "jsonl":
        lines = [json.dumps(r, ensure_ascii=False, default=str)
                 for r in df.to_dict(orient="records")]
        return "\n".join(lines).encode("utf-8"), _MIME["jsonl"]

    if fmt == "xml":
        try:
            out = df.to_xml(index=False, root_name="data", row_name="record")
        except Exception:
            lines = ['<?xml version="1.0" encoding="UTF-8"?>', "<data>"]
            for _, row in df.iterrows():
                lines.append("  <record>")
                for col in df.columns:
                    tag = re.sub(r"[^a-zA-Z0-9_]", "_", str(col))
                    val = "" if row[col] is None else str(row[col]).replace("&", "&amp;").replace("<", "&lt;")
                    lines.append(f"    <{tag}>{val}</{tag}>")
                lines.append("  </record>")
            lines.append("</data>")
            out = "\n".join(lines)
        return out.encode("utf-8"), _MIME["xml"]

    if fmt == "pdf":
        return _to_pdf(df, opt), _MIME["pdf"]

    raise ValueError(f"Formato non supportato: {fmt}")


def _to_pdf(df: pd.DataFrame, opt: dict) -> bytes:
    from reportlab.lib        import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units  import mm
    from reportlab.platypus   import SimpleDocTemplate, Table, TableStyle

    buf  = io.BytesIO()
    page = landscape(A4) if len(df.columns) > 6 else A4
    doc  = SimpleDocTemplate(buf, pagesize=page,
                              leftMargin=10*mm, rightMargin=10*mm,
                              topMargin=15*mm,  bottomMargin=15*mm)

    col_w = (page[0] - 20*mm) / max(len(df.columns), 1)
    data  = [list(df.columns)]
    for _, row in df.iterrows():
        data.append(["" if row[c] is None else str(row[c]) for c in df.columns])

    tbl = Table(data, colWidths=[col_w] * len(df.columns), repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1,  0), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR",      (0, 0), (-1,  0), colors.white),
        ("FONTNAME",       (0, 0), (-1,  0), "Helvetica-Bold"),
        ("FONTNAME",       (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",       (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
        ("GRID",           (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
        ("TOPPADDING",     (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 2),
        ("LEFTPADDING",    (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",   (0, 0), (-1, -1), 4),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
    ]))
    doc.build([tbl])
    return buf.getvalue()


# ── Apply column formatting ───────────────────────────────────────────────────
def apply_formatting(df: pd.DataFrame, col_configs: list[dict]) -> pd.DataFrame:
    """
    col_configs: [{as400_col, label, include, fmt_type, fmt_options, order}]
    Returns a new DataFrame with:
    - only included columns
    - renamed to label
    - formatted
    - ordered
    """
    included = sorted([c for c in col_configs if c.get("include", True)],
                      key=lambda c: c.get("order", 0))

    # Filter cols that exist in df
    existing = {c["as400_col"]: c for c in included if c["as400_col"] in df.columns}
    if not existing:
        return df

    result = df[list(existing.keys())].copy()

    for col, cfg in existing.items():
        fmt_type = cfg.get("fmt_type", "")
        opt      = cfg.get("fmt_options", {}) or {}
        series   = result[col].fillna("")

        if fmt_type == "date" and opt.get("date_format"):
            try:
                parsed = pd.to_datetime(series, errors="coerce")
                result[col] = parsed.dt.strftime(opt["date_format"]).where(parsed.notna(), "")
            except Exception:
                pass

        elif fmt_type == "number":
            dec_sep  = opt.get("decimal_sep",   ".")
            thou_sep = opt.get("thousands_sep",  "")
            try:
                nums = pd.to_numeric(series.str.replace(",", "."), errors="coerce")
                def fmt_num(v):
                    if pd.isna(v): return ""
                    s = f"{v:,.2f}" if thou_sep else f"{v:.2f}"
                    if thou_sep:
                        s = s.replace(",", "THOU").replace(".", dec_sep).replace("THOU", thou_sep)
                    else:
                        s = s.replace(".", dec_sep)
                    return s
                result[col] = nums.apply(fmt_num)
            except Exception:
                pass

        elif fmt_type == "text":
            if opt.get("trim"):      series = series.str.strip()
            if opt.get("uppercase"): series = series.str.upper()
            elif opt.get("lowercase"): series = series.str.lower()
            result[col] = series

    # Rename
    rename_map = {c["as400_col"]: (c.get("label") or c["as400_col"]) for c in existing.values()}
    result     = result.rename(columns=rename_map)

    return result
